import pandas as pd
import numpy as np
import os
import openpyxl
from openpyxl.styles import Border, Side

def _validate_existing(filename):
    filename = _validate_filename(filename);

    if not os.path.exists(filename):
        raise FileNotFoundError("Excel file not found.")
 
    return filename;

def _validate_filename(filename):
    if not filename.endswith('.xlsx'):
        filename += '.xlsx'
    return filename;
    
def read(file_path, remove_header=False):
    """
    Read an Excel file and convert the table contents into a NumPy array.

    Parameters:
    - file_path (str): Relative filepath to the Excel file with or without extension.
    - remove_header (bool): Indicates whether the Excel file has a header. Default is False.

    Returns:
    - np_array (numpy.ndarray): NumPy array containing the table contents.
    """
    try:
        excel_path = _validate_existing(file_path)

        # Load the Excel file using openpyxl
        wb = openpyxl.load_workbook(excel_path, data_only=True)

        # Get the active sheet
        sheet = wb.active

        # Read the data from the sheet into a list of lists
        data = []
        for row in sheet.iter_rows(values_only=True):
            data.append(list(row))

        # If there is a header and it has to be removed
        if remove_header:
            header = data[0]
            data = data[1:]

        # Convert the list of lists into a pandas DataFrame
        df = pd.DataFrame(data)

        # Convert the DataFrame into a NumPy array
        np_array = df.to_numpy()

        return np_array

    except Exception as e:
        print(f"Error occurred: {e}")
        return None

def save(data, filename='output.xlsx'):
    """
    Save the provided data (DataFrame, NumPy array, or dictionary) to an Excel file.

    Parameters:
        data: pandas.DataFrame, numpy.ndarray, or dict
            The data to save. If a NumPy array or dictionary is provided, it will be converted to a DataFrame.
        filename: str
            The name of the Excel file to save the data to.
    """
    try:
        # Convert data to DataFrame if necessary
        if isinstance(data, np.ndarray):
            data = pd.DataFrame(data)
        elif isinstance(data, dict):
            data = pd.DataFrame(data)
        elif not isinstance(data, pd.DataFrame):
            raise ValueError("Input data must be a pandas DataFrame, a numpy ndarray, or a dictionary.")
        
        # Save the DataFrame to an Excel file
        data.to_excel(_validate_filename(filename), index=False, header=True)  # Include headers for dictionaries
    except Exception as e:
        print(f"Error occurred: {e}")
        return None

def add_column_border(filename, first_column=1, last_column=None):
    """
    Adds a border around the specified columns in a worksheet.

    Args:
        ws (Worksheet): The openpyxl worksheet object.
        first_column (int): The index of the first column to add a border (1-based).
        last_column (int): The index of the last column to add a border (1-based).
                           If None, it will default to the last column in the sheet.
    """
    try:
        filename = _validate_existing(filename)

        wb = openpyxl.load_workbook(filename)
        ws = wb.active

        side_border = Border(left=Side(style='thin'), right=Side(style='thin'))

        # Determine the last column if not provided
        if last_column is None:
            last_column = ws.max_column

        # Iterate over the specified columns
        for col_cells in ws.iter_cols(min_col=first_column, max_col=last_column, min_row=1, max_row=ws.max_row):
            for cell in col_cells:
                cell.border = side_border
        
        wb.save(filename)
    except Exception as e:
        print(f"Error occurred: {e}")

def change_column_width(filename, column_width=15, first_column=1, last_column=None):
    """
    Changes the width of specified columns in an existing Excel file.

    Args:
        filename (str): The path to the Excel file.
        column_width (int or float): The width to apply to each column.
        first_column (int): The index of the first column to change width (1-based).
        last_column (int): The index of the last column to change width (1-based).
                           If None, it defaults to the last column in the sheet.
    """
    try:
        # Load the Excel file
        filename = _validate_existing(filename)
        wb = openpyxl.load_workbook(filename)
        ws = wb.active

        # Determine the last column if not provided
        if last_column is None:
            last_column = ws.max_column

        # Iterate over the specified columns and set the width
        for col_idx in range(first_column, last_column + 1):
            col_letter = openpyxl.utils.get_column_letter(col_idx)  # Convert index to letter
            ws.column_dimensions[col_letter].width = column_width

        # Save the modified file
        wb.save(filename)
    except Exception as e:
        print(f"Error occurred: {e}")